import kmedoids
import numpy
from openpyxl import load_workbook
from openpyxl import Workbook
import time
import math
import random
import json

workbook = load_workbook("all_data.xlsx")
worksheet = workbook['output']
all_sentiments = [[] for i in range(34)]
all_text = [[] for i in range(34)]
count = 0
for data_row in worksheet.rows:
    if(count == 0):
        count += 1
        continue
    cluster = int(data_row[5].value)
    if(cluster != -1):
        all_sentiments[cluster].append(json.loads(data_row[3].value))
        all_text[cluster].append(data_row[2].value)

workbook.close()

overall = [[0, 0, 0] for i in range(34)]
#negative, positive, neutral

sent_change = [[0 for i in range(9)] for i in range(34)]
#pos-pos, pos-neu, pos-neg, neu-pos, neu-neu, neu-neg, neg-pos, neg-neu, neg-neg

for i in range(len(all_sentiments)):
    for sentiment in all_sentiments[i]:
        overall[i][0] += sentiment['document']['confidence']['negative']
        overall[i][1] += sentiment['document']['confidence']['positive']
        overall[i][2] += sentiment['document']['confidence']['neutral']
        prev_state = None
        for sent in sentiment['sentences']:
            cur_state = sent['sentiment']
            if(prev_state == None):
                prev_state = sent['sentiment']
                continue
            if(prev_state == 'positive'):
                if(cur_state == 'positive'):
                    sent_change[i][0] += 1
                elif(cur_state == 'neutral'):
                    sent_change[i][1] += 1
                else:
                    sent_change[i][2] += 1
            elif(prev_state == 'neutral'):
                if(cur_state == 'positive'):
                    sent_change[i][3] += 1
                elif(cur_state == 'neutral'):
                    sent_change[i][4] += 1
                else:
                    sent_change[i][5] += 1
            else:
                if(cur_state == 'positive'):
                    sent_change[i][6] += 1
                elif(cur_state == 'neutral'):
                    sent_change[i][7] += 1
                else:
                    sent_change[i][8] += 1
            prev_state = sent['sentiment']
    overall[i][0] = overall[i][0] / len(all_sentiments[i])
    overall[i][1] = overall[i][1] / len(all_sentiments[i])
    overall[i][2] = overall[i][2] / len(all_sentiments[i])

    sent_sum = sum(sent_change[i])
    for j in range(9):
        if(sent_change[i][j] != 0):
            sent_change[i][j] = sent_change[i][j] / sent_sum



print(overall)
print("-----------")
print(sent_change)
        



xlsx = Workbook()
list_sheet = xlsx.create_sheet('output')
list_sheet.append(["negative", "positive", "neutral", "pos to pos", "pos to neu", "pos to neg", "neu to pos", "neu to neu", "neu to neg", "neg to pos", "neg to neu", "neg to neg"])
for i in range(len(overall)):
    list_sheet.append(overall[i] + sent_change[i])

file_name = 'sentiment_result_all.xlsx'
xlsx.save(file_name)


