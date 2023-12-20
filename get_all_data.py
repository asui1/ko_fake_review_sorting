from openpyxl import load_workbook
from openpyxl import Workbook

from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl import Workbook
import torch
import numpy

res_names = ['더스테이크쥬벤쿠바', '어썸로즈', '진작', '고가빈커리하우스', 'H라운지', '살롱순라', '오마', '남산도담', '을지다락여의도', '진대감마포점']
model = SentenceTransformer('snunlp/KR-SBERT-V40K-klueNLI-augSTS') 
data = []
for i in range(len(res_names)):
    dataWorkbook = load_workbook("collected_reviews\\naver_review_" + res_names[i] + ".xlsx")
    dataWorksheet = dataWorkbook['refine11']
    for dataRow in dataWorksheet.rows:
        data.append([res_names[i], dataRow[2].value, dataRow[1].value])
    dataWorkbook.close()
    
for i in range(len(res_names)):
    gpt_workbook = load_workbook('yelp_gpt_kor_data\\gpt_ko_' +res_names[i] +'.xlsx')
    gptWorksheet = gpt_workbook['output']
    for dataRow in gptWorksheet.rows:
        data.append([res_names[i], 2, dataRow[1].value])
    gpt_workbook.close()

xlsx = Workbook()
list_sheet = xlsx.create_sheet('output')
list_sheet.append(["restaurant", "label", "review"])
for i in data:
    list_sheet.append(i)

file_name = 'all_data.xlsx'
xlsx.save(file_name)
print("Finished")
