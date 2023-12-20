

from openpyxl import load_workbook
from openpyxl import Workbook


data = load_workbook("yelp_data\\Yelp_Metadata.xlsx")
worksheet = data['Sheet2']
labels = []
for data_row in worksheet.rows:
    labels.append([data_row[2].value, data_row[3].value])

data.close()



review_data = load_workbook("yelp_data\\Yelp_dataset.xlsx")
worksheet = review_data['Sheet2']
counter = 0
full_data = []
for data_row in worksheet.rows:
    if(labels[counter][1] == -1 and labels[counter][0] > 4 and len(data_row[3].value) > 40):
        full_data.append([data_row[0].value, labels[counter][0], labels[counter][1], data_row[3].value])

    counter += 1
review_data.close

xlsx = Workbook()
list_sheet = xlsx.create_sheet('output')
list_sheet.append(["Restaurant Number", "Rating", "Label", "Review"])
for i in full_data:
    list_sheet.append(i)

file_name = 'yelp_fake.xlsx'
xlsx.save(file_name)


