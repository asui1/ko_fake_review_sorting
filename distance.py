# -*- coding: utf-8 -*-
"""
Created on Thu Oct  5 03:31:30 2023

@author: Asui
"""

from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl import Workbook
import torch
import numpy
from transformers import AutoModel, AutoTokenizer
from transformers import pipeline
import torch


#Mean Pooling - Take attention mask into account for correct averaging
def mean_pooling(model_output, attention_mask):
    token_embeddings = model_output[0] #First element of model_output contains all token embeddings
    input_mask_expanded = attention_mask.unsqueeze(-1).expand(token_embeddings.size()).float()
    return torch.sum(token_embeddings * input_mask_expanded, 1) / torch.clamp(input_mask_expanded.sum(1), min=1e-9)


#The public model snunlp/KR-SBERT-V40K-klueNLI-augSTS (https://github.com/snunlp/KR-SBERT, accessed on 2 May 2023) is based on the KR-BERT-V40K pre-trained model. It was fine-tuned using the KLUE-NLI and KorSTS datasets. The KorSTS dataset was augmented by Augmented SBERT [19];
#The public model jhgan/ko-sroberta-multitask (https://github.com/jhgan00/ko-sentence-transformers, accessed on 2 May 2023) is based on the KLUE-RoBERTa-base pre-trained model. It was fine-tuned using the KorNLI and KorSTS datasets. This model is considered the best among the models introduced on the GitHub page;
#The public model Huffon/sentence-klue-roberta-base (https://huggingface.co/Huffon/sentence-klue-roberta-base, accessed on 2 May 2023) is based on the KLUE-RoBERTa-base pre-trained model. It was fine-tuned using the KLUE-STS dataset.
#goint to use 3 models : snunlp: KR-SBERT, KLUE-RoBERTa-large 모델, MiniLM-L12-v2
#snunlp/KR-SBERT-V40K-klueNLI-augSTS -> cluster 34, randomstate 2 -> SBERT 모델 0.8628
#KoELECTRA-Base-v3 -> KorSTS(spearman) 85.53
#KLUE RoBERTa large   https://huggingface.co/klue/roberta-large -> Roberta 모델 93.35

#다언어 모델
#sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2 -> vector space 모델

resNames = ['더스테이크쥬벤쿠바', '어썸로즈', '진작', '고가빈커리하우스', 'H라운지', '살롱순라', '오마', '남산도담', '을지다락여의도', '진대감마포점']


#model = SentenceTransformer('jhgan/ko-sbert-sts')
model = SentenceTransformer('sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2')
#model = SentenceTransformer('snunlp/KR-SBERT-V40K-klueNLI-augSTS') 

workbook = load_workbook("all_data.xlsx")
worksheet = workbook['output']

data = []
count = 0
for dataRow in worksheet.rows:
    if(count == 0):
        count += 1
        continue

    data.append(dataRow[2].value)
workbook.close()

workbook = load_workbook("self_made_review.xlsx")
worksheet = workbook['Sheet1']
made_data = []
for dataRow in worksheet.rows:
    made_data.append(dataRow[0].value)
workbook.close()

vectors = model.encode(data) # encode sentences into vectors

similarities = util.cos_sim(vectors, vectors) # compute similarity between sentence vectors
similarities_numpy = similarities.numpy()
numpy.save("augsts_numpy_save", similarities_numpy)

made_vectors = model.encode(made_data)
test_similarities = util.cos_sim(made_vectors, vectors)
test_similarities_numpy = test_similarities.numpy()
numpy.save("augsts_test_numpy_save", test_similarities_numpy)
