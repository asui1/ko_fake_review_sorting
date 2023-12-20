from openpyxl import load_workbook
# 파일명
resNames = ['더스테이크쥬벤쿠바', '어썸로즈', '진작', '고가빈커리하우스', 'H라운지', '살롱순라', '오마', '남산도담', '을지다락 여의도', '진대감 마포점']

totalCount = 0
originalTotalCount= 0
for resName in resNames:

    file_name = 'naver_review_' + resName + '.xlsx'

    workbook = load_workbook(file_name)

    worksheet = workbook['output']
    data = []

    curCount = 0
    originalCurCount= 0

    for row in worksheet.rows:
        originalTotalCount += 1
        originalCurCount += 1
        if(row[1].value == None):
            continue
        if(len(row[1].value) > 20):
            data.append([row[0].value, row[1].value, row[2].value])
            totalCount += 1
            curCount += 1

    print("Current Sum is : " + resName + " : " + str(totalCount) + " , " + str(curCount))
    print("Original Sum is : " + resName + " : " + str(originalTotalCount) + " , " + str(originalCurCount))

    refined = workbook.create_sheet('refine1')

    for i in data:
        refined.append(i)

    workbook.save(file_name)