import random
from openpyxl import load_workbook
from openpyxl import Workbook


numbers = []
# 2 ~ 16026
while(len(numbers) < 100):
    number = random.randint(1, 16026)
    if(number not in numbers):
        numbers.append(number)

print(numbers)
review_data = load_workbook("yelp_fake.xlsx")
worksheet = review_data['output']
counter = 0
sampled_data = []
for data_row in worksheet.rows:
    if(counter in numbers):
        sampled_data.append(data_row[3].value)
    counter += 1

xlsx = Workbook()
list_sheet = xlsx.create_sheet('output')
list_sheet.append(["assistant_review", "user"])
for i in sampled_data:
    list_sheet.append([i])

file_name = 'yelp_extract.xlsx'
xlsx.save(file_name)


# Could you make a review of ~~~?