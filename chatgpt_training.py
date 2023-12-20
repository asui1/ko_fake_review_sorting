import os
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl import Workbook

client = OpenAI(api_key="")

res_names = ['더스테이크쥬벤쿠바', '어썸로즈', '진작', '고가빈커리하우스', 'H라운지', '살롱순라', '오마', '남산도담', '을지다락여의도', '진대감마포점']
res_data_num = [1004, 1896, 837, 344, 840, 1335, 527, 451, 1047, 935]
res_eng_data = []

res_eng_names = ["\"The Steakhouse Juventus\", a American steak restaurant", "\'Awesome Rose\", a Italian restaurant", "\"JinJak\", a Japanese restaurant", "\"Gagabin Curry House\", a curry restaurant", 
                 "\"H Lounge\", a brunch shop", "\"Salonsunla\", a Italian bar", "\"Oma\", a Italian restaurant", "\"Namsandodam\", a Korean restaurant", "\"The Basement\", a Italian restaurant", "\"JinDaegam Mago-gu\", a Korean beef restaurant"]

for i in range(9, 10):
  xlsx = Workbook()
  list_sheet = xlsx.create_sheet('output')
  for j in range(res_data_num[i]):
    if(j % 100 == 0):
      print(j)
    response = client.chat.completions.create(
  #  model="gpt-3.5-turbo-1106",
      model="ft:gpt-3.5-turbo-1106:personal::",
      temperature=0.7,
      max_tokens=1000,
      messages=[
        {"role": "system", "content": "You have to make good reviews of the restaurant."},
        {"role": "user", "content": "Make a review about " + res_eng_names[i]},
      ]
    )

    list_sheet.append([response.choices[0].message.content])

  file_name = 'yelp_gpt_data\\gpt_' +res_names[i] +'.xlsx'
  xlsx.save(file_name)
  print(res_names[i], " Done")
